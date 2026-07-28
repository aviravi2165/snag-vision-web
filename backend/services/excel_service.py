"""
Activity Excel parsing + in-place progress sync.

The uploaded Activity Excel is the project's master progress sheet. We never
regenerate it from a template — we open the SAME file the user uploaded, write
only the progress cells that have real AI-derived data, and save it back with
every formula, other sheet, color, and data validation untouched.

Header/room-column detection mirrors the same flexible heuristics already
proven client-side in frontend/src/pages/Projects.jsx::handleExcelUpload —
ported here so activity names and room columns are never hardcoded per
project, and the two code paths agree on what "looks like" a name/date column.
"""
import re
from dataclasses import dataclass, field
from datetime import datetime, date
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl

NAME_COL_RE = re.compile(r"activity|task|item|scope|description|work", re.I)
DATE_COL_RE = re.compile(r"date|deadline|target|completion|due", re.I)
HEADER_SKIP_RE = re.compile(r"^activity(\s*(name|plan))?$", re.I)
# A trailing summary/formula column (e.g. "Overall Progress %", "Average", "Total") —
# not a Room at all, so it must never be offered up for manual room-linking.
SUMMARY_COL_RE = re.compile(r"overall|average|avg|total|summary", re.I)


@dataclass
class ParsedActivity:
    name: str
    start_date: Optional[datetime]
    end_date: Optional[datetime]


@dataclass
class ParsedUnitColumn:
    """A column in the Activity Excel representing one Unit (the business
    "Room ID", e.g. "A-101") — the Excel's actual room-column grain. Not to
    be confused with the app's `Room` model, which is an Area (Living/
    Bathroom) *within* a Unit — Area-level breakdown is Floor View's job."""
    col_index: int  # 1-based openpyxl column index
    header: str


@dataclass
class ParsedExcel:
    sheet_name: str
    header_row: int
    activity_col_index: int
    activity_col: str
    start_date_col: Optional[str]
    end_date_col: Optional[str]
    activities: List[ParsedActivity] = field(default_factory=list)
    unit_columns: List[ParsedUnitColumn] = field(default_factory=list)


def _is_numeric_column(values: List) -> bool:
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, (datetime, date)):
            return False
        try:
            float(v)
        except (TypeError, ValueError):
            return False
    return True


def _to_datetime(v) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt)
        except ValueError:
            continue
    return None


def parse_activity_excel(file_bytes: bytes) -> ParsedExcel:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header_row = 1
    max_col = ws.max_column
    max_row = ws.max_row

    headers = {
        c: (str(ws.cell(row=header_row, column=c).value).strip()
            if ws.cell(row=header_row, column=c).value is not None else "")
        for c in range(1, max_col + 1)
    }
    col_values = {
        c: [ws.cell(row=r, column=c).value for r in range(header_row + 1, max_row + 1)]
        for c in range(1, max_col + 1)
    }

    # Date columns (start, optionally end) — matched by header keyword.
    date_col_idx = None
    end_date_col_idx = None
    for c in range(1, max_col + 1):
        if DATE_COL_RE.search(headers[c]):
            if date_col_idx is None:
                date_col_idx = c
            elif end_date_col_idx is None:
                end_date_col_idx = c

    # Activity-name column — prefer an explicitly-labelled, non-numeric column;
    # fall back to the first non-numeric, non-date column (skips serial-number
    # columns like "S.No"/"#" the same way the frontend parser does).
    name_col_idx = None
    for c in range(1, max_col + 1):
        if c in (date_col_idx, end_date_col_idx):
            continue
        if NAME_COL_RE.search(headers[c]) and not _is_numeric_column(col_values[c]):
            name_col_idx = c
            break
    if name_col_idx is None:
        for c in range(1, max_col + 1):
            if c in (date_col_idx, end_date_col_idx):
                continue
            if not _is_numeric_column(col_values[c]):
                name_col_idx = c
                break
    if name_col_idx is None:
        name_col_idx = 1

    # Every column after the last identity (name/date) column, with a non-blank
    # header, is a Unit column (the Excel's "Room" grain — e.g. "A-101") —
    # except a trailing summary/formula column (e.g. "Overall Progress %"),
    # which isn't a Room at all and must never be offered up for linking.
    identity_cols = {c for c in (name_col_idx, date_col_idx, end_date_col_idx) if c}
    last_identity_col = max(identity_cols) if identity_cols else name_col_idx
    unit_columns = [
        ParsedUnitColumn(col_index=c, header=headers[c])
        for c in range(last_identity_col + 1, max_col + 1)
        if headers[c] and not SUMMARY_COL_RE.search(headers[c])
    ]

    activities: List[ParsedActivity] = []
    for r in range(header_row + 1, max_row + 1):
        raw_name = ws.cell(row=r, column=name_col_idx).value
        name = str(raw_name).strip() if raw_name is not None else ""
        if not name or HEADER_SKIP_RE.match(name):
            continue
        start_date = _to_datetime(ws.cell(row=r, column=date_col_idx).value) if date_col_idx else None
        end_date = _to_datetime(ws.cell(row=r, column=end_date_col_idx).value) if end_date_col_idx else None
        activities.append(ParsedActivity(name=name, start_date=start_date, end_date=end_date))

    return ParsedExcel(
        sheet_name=ws.title,
        header_row=header_row,
        activity_col_index=name_col_idx,
        activity_col=headers[name_col_idx],
        start_date_col=headers.get(date_col_idx) if date_col_idx else None,
        end_date_col=headers.get(end_date_col_idx) if end_date_col_idx else None,
        activities=activities,
        unit_columns=unit_columns,
    )


def match_unit_columns(
    unit_columns: List[ParsedUnitColumn], db_units: List
) -> Tuple[Dict[str, int], List[ParsedUnitColumn]]:
    """Match Excel column headers to real Unit.unit_number rows created in
    Setup (the "Room ID" business concept, e.g. "A-101") by normalized
    equality only — never a guess. Also accepts a "Floor N | A-101"-style
    composite header (as produced by this app's own CSV export) by matching
    on the part after the last "|". Unit numbers are unique per project, so
    unlike Area names (Living/Bathroom, which repeat across units) this
    should resolve unambiguously in the normal case — but if a header still
    matches more than one Unit, that's surfaced for manual linking rather
    than guessed. Returns ({unit_id: col_index} for unambiguous matches,
    [unmatched/ambiguous columns])."""
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", s.strip().lower())

    units_by_number: Dict[str, List] = {}
    for u in db_units:
        units_by_number.setdefault(norm(u.unit_number), []).append(u)

    matched: Dict[str, int] = {}
    unmatched: List[ParsedUnitColumn] = []
    for col in unit_columns:
        header = col.header
        candidates = units_by_number.get(norm(header), [])
        if not candidates and "|" in header:
            candidates = units_by_number.get(norm(header.rsplit("|", 1)[-1]), [])
        if len(candidates) == 1:
            matched[candidates[0].id] = col.col_index
        else:
            unmatched.append(col)  # 0 matches (unknown) or >1 (ambiguous) — either way, don't guess
    return matched, unmatched


def sync_excel_progress(
    file_bytes: bytes,
    header_row: int,
    activity_col_index: int,
    activities: List,  # models.database.Activity rows (id, name)
    unit_col_map: Dict[str, int],  # {unit_id: col_index}
    progress_by_activity_unit: Dict[Tuple[str, str], Optional[float]],  # {(activity_id, unit_id): pct}
) -> bytes:
    """Open the master workbook in place and write only the cells that have a
    real progress value. Everything else — formulas, other sheets, colors,
    validations, untouched activities/units — is left exactly as-is."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)
    ws = wb.active
    max_row = ws.max_row

    name_to_id = {a.name.strip().lower(): a.id for a in activities}

    for r in range(header_row + 1, max_row + 1):
        raw_name = ws.cell(row=r, column=activity_col_index).value
        if raw_name is None:
            continue
        activity_id = name_to_id.get(str(raw_name).strip().lower())
        if not activity_id:
            continue
        for unit_id, col_idx in unit_col_map.items():
            pct = progress_by_activity_unit.get((activity_id, unit_id))
            if pct is None:
                continue  # never assessed yet — leave the cell as the uploaded sheet had it
            cell = ws.cell(row=r, column=col_idx)
            fmt = cell.number_format or ""
            if "%" in fmt:
                cell.value = round(pct, 2) / 100.0
            else:
                cell.value = f"{round(pct)}%"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
